#!/usr/bin/env python
# -*- coding: utf8 -*-

import yfinance as yf
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl.chart import LineChart, Reference
import sys
from datetime import datetime, timedelta
import os
import pdfkit
config = pdfkit.configuration(wkhtmltopdf='c:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe')

from fpdf import FPDF

# 0. GET SAMPLE DATA
sp500 = yf.Ticker("^GSPC")
end_date = pd.Timestamp.today()
start_date = end_date - pd.Timedelta(days=10 * 365)
sp500_history = sp500.history(start=start_date, end=end_date)

sp500_history = sp500_history.drop(columns=['Dividends', 'Stock Splits'])
sp500_history['Close_200ma'] = sp500_history['Close'].rolling(200).mean()
sp500_history_summary = sp500_history.describe()

# 0-1. SAVE CHART
sns.relplot(data=sp500_history[['Close', 'Close_200ma']], kind='line', height=3, aspect=2.0)
if not os.path.isfile('~/chart.png'):
    plt.savefig('chart.png')


class TestHandler:

    def __init__(self):
        self.type = TYPE

    def run(self):

        testType = self.type

        if testType == 'e':
            self.excel()
        elif testType == 'h':
            self.html()
        elif testType == 'p':
            # self.direct()
            pdf = FPDF(orientation='P', unit='mm', format='A4')
            pdf.add_page()
            pdf.add_font('malgun', '', 'malgun.ttf', uni=True)
            self.generate_pdf(pdf)
            pdf.output('unicode.pdf', 'F')

    def excel(self):

        # 1. EXCEL
        with pd.ExcelWriter('[1]excel_report.xlsx', engine='openpyxl') as writer:
            sp500_history.to_excel(writer, sheet_name='historical_data')
            sp500_history_summary.to_excel(writer, sheet_name='hist_summary_data')

            # 1-1. ADD LINECHART
            wb = writer.book
            ws = wb['historical_data']
            max_row = ws.max_row  # Grab the maximum row number in the sheet
            values_close = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=max_row)
            values_close_ma = Reference(ws, min_col=7, min_row=1, max_col=7, max_row=max_row)
            dates = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=max_row)

            # 1-2. CREATE LINECHART
            chart = LineChart()
            chart.add_data(values_close, titles_from_data=True)
            chart.add_data(values_close_ma, titles_from_data=True)
            chart.set_categories(dates)
            chart.x_axis.number_format = 'mmm-yy'
            chart.x_axis.majorTimeUnit = 'days'
            chart.x_axis.title = 'Date'

            chart.title = 'Close prices of S&P 500'
            # Refer to close_ma data, which is with index 1 within the chart, and style it
            s1 = chart.series[1]
            s1.graphicalProperties.line.dashStyle = 'sysDot'
            ws.add_chart(chart, 'G12')

    def html(self):

        html = f'''
                <html>
                    <head>
                        <meta charset="utf-8">
                        <title>일일 점검 리포트</title>
                    </head>
                    <body>
                        <h1>일일 점검 리포트</h1>
                        <h2>1.진단 결과</h2>
                        <img src='res/inspection.png' width="250">
                        <img src='res/inspection.png' width="250">
                        <img src='res/inspection.png' width="250">
                        <h2>2. to-do list</h2>
                        <br>
                        <h2>3.일주일 내 비정상 발생 건수</h2>
                        {sp500_history.tail(5).to_html()}
                        <h2>4. 상세 정보</h2>
                        {sp500_history_summary.to_html()}
                    </body>
                </html>
                '''

        with open('htmlReport.html', 'w', encoding='UTF-8') as f:
            f.write(html)

        pdfkit.from_file('htmlReport.html', 'htmlToPdf.pdf', configuration=config)

    def output_df_to_pdf(self, pdf, df):

        table_cell_width = 25
        table_cell_height = 6
        pdf.set_font('Arial', 'B', 8)

        cols = df.columns
        for col in cols:
            pdf.cell(table_cell_width, table_cell_height, col, align='C', border=1)
        pdf.ln(table_cell_height)
        pdf.set_font('Arial', '', 10)
        for row in df.itertuples():
            for col in cols:
                value = str(getattr(row, col))
                pdf.cell(table_cell_width, table_cell_height, value, align='C', border=1)
            pdf.ln(table_cell_height)

    def write_to_pdf(self, pdf, words):

        pdf.set_text_color(r=0, g=0, b=0)
        pdf.set_font('Arial', '', 12)
        pdf.write(5, words)

    def write_title(self, pdf, txt):

        pdf.set_text_color(r=0, g=0, b=0)
        pdf.set_font('malgun', '', 12)
        pdf.write(5, txt)

    def direct(self):

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font('Arial', 'B', 16)

        pdf.cell(40, 10, 'Daily S&P 500 prices report')
        pdf.ln(10)

        d = datetime.now()
        yd = d - timedelta(1)
        insp_date = f'inspection date: {yd} ~ {d}'
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(40, 10, insp_date)
        pdf.ln(10)

        s = '1. 형성 점검 결과'.encode('utf-8').decode('latin-1')
        pdf.cell(40, 10, s)
        pdf.ln(30)
        pdf.image('myplot.png', 15, 40, 30, 20)

        s = '2. 자동 점검 결과'.encode('UTF-8').decode('latin-1')
        pdf.cell(40, 10, s)
        pdf.image('myplotx.png', 15, 70, 45)
        pdf.image('chart.png', 65, 70, 45)
        pdf.image('chart.png', 110, 70, 45)
        pdf.image('chart.png', 155, 70, 45)

        sp500_history_pdf = sp500_history.reset_index()
        sp500_history_pdf['Date'] = sp500_history_pdf['Date'].astype(str)
        numeric_cols = sp500_history_pdf.select_dtypes(include='number').columns
        sp500_history_pdf[numeric_cols] = sp500_history_pdf[numeric_cols].round(2)
        pdf.ln(50)

        pdf.set_font('Arial', 'B', 12)
        pdf.cell(40, 10, '3. to-do list')

        pdf.ln(30)
        pdf.cell(40, 10, '4. Fail Count Data')
        pdf.ln(10)
        self.output_df_to_pdf(pdf, sp500_history_pdf.tail(5))

        pdf.ln(10)
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(40, 10, '5. Detaled Fail Data')
        pdf.ln(10)

        sp500_history_summary_pdf = sp500_history_summary.reset_index()
        numeric_cols = sp500_history_summary_pdf.select_dtypes(include='number').columns
        sp500_history_summary_pdf[numeric_cols] = sp500_history_summary_pdf[numeric_cols].round(2)
        self.output_df_to_pdf(pdf, sp500_history_summary_pdf)
        pdf.output('[3]pdf_direct.pdf', 'F')

    def generate_pdf(self, pdf):

        # HEADER
        pdf.set_font('malgun', '', 14)
        pdf.cell(self.WIDTH - 130)
        pdf.write(8, u'일일 점검 리포트')
        pdf.ln(10)

        # Date
        d = datetime.now()
        yd = d - timedelta(1)
        insp_date_str = u'점검 날짜:'
        pdf.set_font('malgun', '', 8)
        pdf.write(5, f'{insp_date_str}: {yd} ~ {d}')
        pdf.ln(10)

        # BODY
        self.write_title(pdf, u'1. 형성 점검 결과')
        pdf.image('res/inspection.png', 5, 35, 70, 60)
        pdf.image('res/inspection.png', 70, 35, 70, 60)
        pdf.image('res/inspection.png', 140, 35, 70, 60)
        pdf.ln(65)

        self.write_title(pdf, u'2. 자동 점검 결과')
        pdf.ln(120)

        self.write_title(pdf, u'3. to-do list')
        pdf.ln(30)

        self.write_title(pdf, u'4. 일주일 내 비정상 발생 건수')
        pdf.ln(20)

        self.write_title(pdf, u'5. 비정상 점검 결과 상세 정보')
        pdf.ln(10)


if __name__ == "__main__":

    while True:
        TYPE = input('[e] excel [h] html [p] pdf : ')
        test = TestHandler()
        test.run()
