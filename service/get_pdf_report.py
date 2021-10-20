#!/usr/bin/env python
# -*- coding: utf8 -*-

import yfinance as yf
import numpy as np
import pandas as pd
from fpdf import FPDF

import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.ticker import ScalarFormatter

import seaborn as sns
from openpyxl.chart import LineChart, Reference
from datetime import datetime, timedelta
import pdfkit
config = pdfkit.configuration(wkhtmltopdf='c:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe')
import calendar


WIDTH = 210
HEIGHT = 297

# 0. GET SAMPLE DATA
sp500 = yf.Ticker("^GSPC")
end_date = pd.Timestamp.today()
start_date = end_date - pd.Timedelta(days=10 * 365)
sp500_history = sp500.history(start=start_date, end=end_date)

sp500_history = sp500_history.drop(columns=['Dividends', 'Stock Splits'])
sp500_history['Close_200ma'] = sp500_history['Close'].rolling(200).mean()
sp500_history_summary = sp500_history.describe()

# 0-1. SAVE CHART
# sns.relplot(data=sp500_history[['Close', 'Close_200ma']], kind='line', height=3, aspect=2.0, legend='d')
# if not os.path.isfile('/chart.png'):
#     plt.savefig('chart.png')

def plot1(filename:str) -> None:

    plt.figure(figsize=(12, 4))

    x = np.arange(3)
    years = ['CPU 사용률', '팬상태', '포트상태']
    values = [500, 400, 100]

    plt.plot(years, values)

    plt.bar(x, values, width=0.4, color=['r', 'g', 'b'])
    plt.xticks(x, years)

    plt.savefig(filename)


    return


def plot2(filename:str) -> None:

    plt.figure(figsize=(12, 4))

    y = np.arange(3)
    years = ['CPU', 'FAN', 'PORT']
    values = [100, 400, 900]

    plt.barh(y, values)
    plt.yticks(y, years)

    plt.show()
    plt.savefig(filename)
    plt.close()

    return

class PdfGenerator:
    def __init__(self):
        self.WIDTH = WIDTH
        self.HEIGHT = HEIGHT

    def run(self):

        self.direct()

    def output_df_to_pdf(self, pdf, df):

        table_cell_width = 25
        table_cell_height = 6

        pdf.set_font('malgun', '', 6)


        # Loop over to print column names
        cols = df.columns
        for col in cols:
            pdf.cell(table_cell_width, table_cell_height, col, align='C', border=1)

        pdf.ln(table_cell_height)

        # Loop over to print each data in the table
        for row in df.itertuples():
            for col in cols:
                value = str(getattr(row, col))
                pdf.cell(table_cell_width, table_cell_height, value, align='C', border=1)
            pdf.ln(table_cell_height)

    def write_to_pdf(self,pdf, words):

        pdf.set_text_color(r=0, g=0, b=0)
        pdf.set_font('Arial', '', 12)
        pdf.write(5, words)

    def generate_sales_data(month: int) -> pd.DataFrame:
        # Date range from first day of month until last
        # Use ```calendar.monthrange(year, month)``` to get the last date
        dates = pd.date_range(
            start=datetime(year=2021, month=month, day=1),
            end=datetime(year=2021, month=month, day=calendar.monthrange(2021, month)[1])
        )

        # Sales numbers as a random integer between 1000 and 2000
        sales = np.random.randint(low=1000, high=2000, size=len(dates))

        # Combine into a single dataframe
        return pd.DataFrame({
            'Date': dates,
            'ItemsSold': sales
        })

    def chart_hist(self):
        pass



    def direct(self):

        # 1. Set up the PDF doc basics
        pdf = FPDF()
        pdf.add_page()
        pdf.add_font('malgun', '', 'malgun.ttf', uni=True)

        # 2. Layout the PDF doc contents
        ## Title
        pdf.cell(70)
        pdf.write(8, u'일일 점검 리포트')
        pdf.ln(10)

        d = datetime.now()
        yd = d - timedelta(1)
        insp_date_str = u'점검 날짜:'
        insp_date = f'{insp_date_str}: {yd} ~ {d}'
        pdf.set_font('malgun', '', 8)
        pdf.write(8, insp_date)
        pdf.ln(5)

        pdf.set_font('malgun', '', 10)
        pdf.write(8, u'1. 형성 점검 결과')
        pdf.ln(20)

        pdf.write(8, u'2. 자동 점검 결과')
        pdf.image('myplot.png', 5, 55, 50, 30)
        pdf.image('myplot.png', 110, 55, 50, 30)
        # 줄바꿈
        pdf.image('myplotx.png', 5, 100, 40, 40)
        pdf.image('savefig_default.png', 65, 100, 45, 45)
        pdf.image('chart.png', 110, 100, 45, 45)
        pdf.image('chart.png', 155, 100, 45, 45)


        ## Show table of historical data
        ### Transform the DataFrame to include index of Date
        sp500_history_pdf = sp500_history.reset_index()
        ### Transform the Date column as str dtype
        sp500_history_pdf['Date'] = sp500_history_pdf['Date'].astype(str) # date col -> str
        ### Round the numeric columns to 2 decimals
        numeric_cols = sp500_history_pdf.select_dtypes(include='number').columns
        sp500_history_pdf[numeric_cols] = sp500_history_pdf[numeric_cols].round(2)
        ### Use the function defined earlier to print the DataFrame as a table on the PDF

        pdf.ln(100)

        pdf.write(8, u'3. to-do list')

        pdf.ln(30)
        pdf.write(8, u'4. 일주일 내 비정상 발생 건수')
        pdf.ln(10)
        self.output_df_to_pdf(pdf, sp500_history_pdf.tail(5))

        pdf.ln(5)
        pdf.set_font('malgun', '', 10)
        pdf.write(8, u'5. 비정상 점검 결과 상세 정보')
        pdf.ln(10)

        sp500_history_summary_pdf = sp500_history_summary.reset_index()
        numeric_cols = sp500_history_summary_pdf.select_dtypes(include='number').columns
        sp500_history_summary_pdf[numeric_cols] = sp500_history_summary_pdf[numeric_cols].round(2)
        self.output_df_to_pdf(pdf, sp500_history_summary_pdf.tail(4))

        # 3. Output the PDF file
        pdf.output('INSPECTION_REPORT.pdf', 'F')


if __name__ == "__main__":

    plot1(filename='xstick.png')
    plot2(filename='ystick.png')

    test = PdfGenerator()

    test.run()
